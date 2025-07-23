from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime
from flask import current_app
import os


def get_monthly_records(filename: str, sheet_name: str) -> dict:
    """
    Excelファイルから月次のレコードを取得する

    Parameters:
    filename (str): Excelファイルのパス
    sheet_name (str): 対象のシート名

    Returns:
    dict: 月ごとのレコードを格納した辞書
    """

    wb = load_workbook(filename)
    sheet = wb[sheet_name]

    monthly = defaultdict(list)
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        vals = [c.value for c in row]
        if not vals[0]:
            continue
        try:
            date_obj = datetime.strptime(vals[0], "%Y-%m-%d")
        except ValueError:
            continue

        reported = vals[7] in [True, "TRUE", "true", 1, "1"] if len(vals) > 7 else False
        month_key = date_obj.strftime("%Y-%m")
        monthly[month_key].append({
            "row": idx,
            "date": vals[0],
            "date_obj": date_obj,
            "start": vals[1],
            "end": vals[2],
            "break": vals[3],
            "worked": vals[4],
            "project": vals[5],
            "content": vals[6],
            "reported": reported
        })

    for recs in monthly.values():
        recs.sort(key=lambda x: x["date_obj"])

    return dict(sorted(monthly.items()))


def update_reported_flags(form_data):
    """
    入力フォームからのデータを基に、Excelファイルの報告フラグを更新する

    Parameters:
    form_data (dict): 入力フォームからのデータ
    """

    path = os.path.join(current_app.instance_path, current_app.config["EXCEL_FILENAME"])
    wb = load_workbook(path)
    sheet = wb[current_app.config["EXCEL_SHEET"]]
    for row in range(2, sheet.max_row + 1):
        key = f"reported_{row}"
        sheet.cell(row=row, column=8).value = key in form_data
    wb.save(path)
