import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from flask import current_app


def ensure_excel_initialized():
    """
    Excelファイルが存在しない場合、初期化する

    """
    path = os.path.join(current_app.instance_path, current_app.config["EXCEL_FILENAME"])
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = current_app.config["EXCEL_SHEET"]
        write_header(ws)
        wb.save(path)


def write_header(sheet):
    """
    Excelシートにヘッダーを書き込む

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): ヘッダーを書き込むシート
    """
    headers = ["日付", "出勤時間", "退勤時間", "休憩時間", "実働時間", "プロジェクト", "内容", "報告"]
    for i, h in enumerate(headers, 1):
        sheet.cell(row=1, column=i, value=h)


def save_record(form):
    """
    入力フォームからのデータをExcelファイルに保存する

    Parameters:
    form (dict): 入力フォームからのデータ
    """

    path = os.path.join(current_app.instance_path, current_app.config["EXCEL_FILENAME"])
    wb = load_workbook(path)
    sheet = wb[current_app.config["EXCEL_SHEET"]]

    # 入力データの検証
    date_str = form["date"]
    start = f"{form['start_hour']}:{form['start_minute']}"
    end = f"{form['end_hour']}:{form['end_minute']}"
    break_time = form.get("break_time", "00:00").strip()

    # 休憩時間が空または不正な場合、デフォルト値を設定
    if not break_time or ":" not in break_time:
        break_time = "00:00"

    start_dt = datetime.strptime(f"{date_str} {start}", "%Y-%m-%d %H:%M")
    end_dt = datetime.strptime(f"{date_str} {end}", "%Y-%m-%d %H:%M")
    # 退勤時間が出勤時間より前の場合、翌日に設定
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)

    # 休憩時間を分解して計算
    h, m = map(int, break_time.split(":"))
    break_delta = timedelta(hours=h, minutes=m)

    # 実働時間を計算
    worked = end_dt - start_dt - break_delta
    worked_str = f"{int(worked.total_seconds()//3600):02d}:{int((worked.total_seconds()%3600)//60):02d}"

    # 新しい行を追加
    sheet.append([
        date_str, start, end,
        break_time, worked_str,
        form["project"], form["content"], False
    ])
    wb.save(path)
