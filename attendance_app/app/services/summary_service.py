from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime, timedelta


def parse_duration(hhmm: str) -> timedelta:
    """
    文字列形式の時間（HH:MM）をtimedeltaに変換する

    Parameters:
    hhmm (str): 時間を表す文字列（例: "01:30"）

    Returns:
    timedelta: 時間を表すtimedeltaオブジェクト
    """
    h, m = map(int, hhmm.strip().split(":"))
    return timedelta(hours=h, minutes=m)


def get_monthly_summary(filename: str, sheet_name: str, hourly_rate: int, target_salary: int) -> list:
    """
    Excelファイルから月次のサマリを取得する

    Parameters:
    filename (str): Excelファイルのパス
    sheet_name (str): 対象のシート名
    hourly_rate (int): 時給
    target_salary (int): 目標給与

    Returns:
    list: 月次のサマリデータのリスト
    """

    wb = load_workbook(filename)
    sheet = wb[sheet_name]
    summary = defaultdict(lambda: {"total_time": timedelta(), "date_set": set(), "unreported": 0})

    for row in sheet.iter_rows(min_row=2, values_only=True):
        date_str, _, _, _, dur, _, _, rep = row
        if not date_str or not dur:
            continue
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            dur_td = parse_duration(dur)
        except Exception:
            continue

        key = dt.strftime("%Y-%m")
        summary[key]["total_time"] += dur_td
        summary[key]["date_set"].add(dt.date())
        summary[key]["unreported"] += 0 if rep else 1

    results = []
    for month, data in sorted(summary.items()):
        hours = data["total_time"].total_seconds() / 3600
        salary = int(hours * hourly_rate)
        remain = target_salary - salary
        results.append({
            "month": month,
            "days": len(data["date_set"]),
            "hours": round(hours, 2),
            "salary": salary,
            "needed_hours": round(max(0, remain / hourly_rate), 2),
            "unreported": data["unreported"]
        })

    return results
